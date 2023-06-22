import pandas as pd
import re
from openpyxl import load_workbook

# def data():
    # excel data from https://bitbucket.org/dpat/tools/src/42eb353fdf29/REF/ATLASES/Glasser_2016_Table.xlsx?at=master
    path = "/home/kiwitz1/PycharmProjects/openMINDS_SANDS/helper_code/Glasser_2016_Table.xlsx"

    # children
    data = pd.read_excel(path, header=1)
    # clean data
    areas = list(data["Area Description"])
    removed_brackets_data = [re.sub(r'\([^)]*\)', '', string) for string in areas]
    replaced_data = [string.replace("\n", " ") for string in removed_brackets_data]

    # USE CHAT GPT SOLUTION  https://chat.openai.com/share/1c5fdb63-dbf6-49a1-9b88-2af7020f2a8d
    # parents
    parents = list(data["Sections"])
    workbook = load_workbook(path)
    sheet = workbook.active
    bold_content_list = []
    # Iterate over the rows in the "Sections" column, starting from the third row
    for row in sheet.iter_rows(min_row=3):

        cell = row[4]  # Assuming the "Sections" column is in the second column (index 1)
        print(cell)
        if cell.font.bold and cell.value is not None:
            # Convert the cell value to string
            cell_value_str = str(cell.value)

            # Split the comma-separated entries
            entries = cell_value_str.split(",")

            # Filter and append the bold entries
            bold_entries = [entry.strip() for entry in entries if sheet.cell(row=row[0].row, column=2).font.bold]
            bold_content_list.extend(bold_entries)

# Print the extracted bold content
print(bold_content_list)

    def to_lower_camel_case(string):
        words = string.split()  # Split the string into words
        # Convert the first word to lowercase and the rest to title case
        words = [words[0].lower()] + [word.title() for word in words[1:]]
        return ''.join(words)  # Join the words back together


    lower_camel_case_data = [to_lower_camel_case(string) for string in replaced_data]

    # return lower_camel_case_data

# if __name__ == '__main__':

  #  areas = data()